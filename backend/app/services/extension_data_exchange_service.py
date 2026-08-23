from __future__ import annotations
import base64,csv,hashlib,io,ipaddress,json,socket,urllib.error,urllib.parse,urllib.request
from typing import Any

class ExtensionDataExchangeService:
    MAX_BYTES=5*1024*1024;MAX_ROWS=10_000;MAX_COLUMNS=200
    def __init__(self,repository:Any)->None:self.repository=repository
    def stage_text(self,*,name:str,filename:str,content:str,actor:str,source_type:str="file",source_uri:str|None=None)->dict[str,Any]:
        raw=content.encode('utf-8')
        if len(raw)>self.MAX_BYTES:raise ValueError("扩展文件超过 5MB")
        suffix=filename.lower().rsplit('.',1)[-1] if '.' in filename else ''
        if suffix=='csv':
            reader=csv.DictReader(io.StringIO(content));columns=list(reader.fieldnames or []);rows=[dict(row) for row in reader]
        elif suffix=='json':
            payload=json.loads(content);rows=payload if isinstance(payload,list) else [payload];columns=sorted({str(key) for row in rows if isinstance(row,dict) for key in row})
            if not all(isinstance(row,dict) for row in rows):raise ValueError("JSON 必须是对象或对象数组")
        else:raise ValueError("只接受 CSV 或 JSON 文本暂存")
        if not columns or len(columns)>self.MAX_COLUMNS:raise ValueError("扩展数据列数无效或超过 200")
        if len(rows)>self.MAX_ROWS:raise ValueError("扩展数据超过 10000 行")
        for row in rows:
            for value in row.values():
                text=str(value or '').lstrip()
                if text.startswith(('=','+','@')) or (text.startswith('-') and not self._numeric(text)):raise ValueError("扩展数据包含公式或公式注入前缀")
        result=self.repository.stage(name=name,filename=filename,file_format=suffix,content_hash=hashlib.sha256(raw).hexdigest(),size_bytes=len(raw),columns=columns,rows=rows,actor=actor,source_type=source_type,source_uri=source_uri)
        return {**result,"mapping_state":"staged_only","execution_eligible":False}
    def stage_xlsx(self,*,name:str,filename:str,content_base64:str,actor:str)->dict[str,Any]:
        raw=base64.b64decode(content_base64,validate=True)
        if len(raw)>self.MAX_BYTES:raise ValueError("扩展文件超过 5MB")
        from openpyxl import load_workbook
        workbook=load_workbook(io.BytesIO(raw),read_only=True,data_only=False);sheet=workbook.active;iterator=sheet.iter_rows(values_only=True)
        try:columns=[str(value or '').strip() for value in next(iterator)]
        except StopIteration:raise ValueError("XLSX 为空")
        if not columns or len(columns)>self.MAX_COLUMNS or any(not value for value in columns):raise ValueError("XLSX 列名无效或超过 200")
        rows=[]
        for values in iterator:
            if len(rows)>=self.MAX_ROWS:raise ValueError("扩展数据超过 10000 行")
            row=dict(zip(columns,values));self._reject_formulas(row);rows.append(row)
        result=self.repository.stage(name=name,filename=filename,file_format='xlsx',content_hash=hashlib.sha256(raw).hexdigest(),size_bytes=len(raw),columns=columns,rows=rows,actor=actor)
        return {**result,"mapping_state":"staged_only","execution_eligible":False}
    def stage_http(self,*,name:str,url:str,allowlist:set[str],actor:str)->dict[str,Any]:
        parsed=urllib.parse.urlparse(url);host=(parsed.hostname or '').lower()
        if parsed.scheme!='https' or host not in allowlist:raise ValueError("HTTPS 来源不在精确 allowlist")
        addresses={item[4][0] for item in socket.getaddrinfo(host,443,type=socket.SOCK_STREAM)}
        if not addresses or any(not ipaddress.ip_address(address).is_global for address in addresses):raise ValueError("HTTPS 来源必须解析到公共地址")
        class NoRedirect(urllib.request.HTTPRedirectHandler):
            def redirect_request(self,*args,**kwargs):return None
        request=urllib.request.Request(url,headers={'User-Agent':'StockPro-Extension-Import/1'})
        try:
            with urllib.request.build_opener(NoRedirect).open(request,timeout=15) as response:raw=response.read(self.MAX_BYTES+1)
        except urllib.error.HTTPError as error:raise ValueError(f"HTTPS 来源拒绝或重定向: {error.code}") from error
        if len(raw)>self.MAX_BYTES:raise ValueError("HTTPS 扩展文件超过 5MB")
        filename=urllib.parse.unquote(parsed.path.rsplit('/',1)[-1]) or 'import.csv';content=raw.decode('utf-8-sig');result=self.stage_text(name=name,filename=filename,content=content,actor=actor,source_type='http',source_uri=url)
        return {**result,"source_type":"http","source_uri":url}
    def export(self,import_id:str,file_format:str)->tuple[str,str,str]:
        records=self.repository.import_records(import_id);rows=[dict(item['payload']) for item in records]
        if file_format=='json':return json.dumps(rows,ensure_ascii=False,default=str),'application/json',f'{import_id}.json'
        columns=sorted({key for row in rows for key in row});output=io.StringIO();writer=csv.DictWriter(output,fieldnames=columns);writer.writeheader()
        for row in rows:writer.writerow({key:self._safe_export(value) for key,value in row.items()})
        return output.getvalue(),'text/csv; charset=utf-8',f'{import_id}.csv'
    def _reject_formulas(self,row:dict[str,Any])->None:
        for value in row.values():
            text=str(value or '').lstrip()
            if text.startswith(('=','+','@')) or (text.startswith('-') and not self._numeric(text)):raise ValueError("扩展数据包含公式或公式注入前缀")
    @staticmethod
    def _safe_export(value:Any)->Any:
        text=str(value or '')
        return "'"+text if text.lstrip().startswith(('=','+','-','@')) else value
    @staticmethod
    def _numeric(value:str)->bool:
        try:float(value);return True
        except ValueError:return False
