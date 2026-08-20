import io
import json
import unittest
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from app.services.extension_data_exchange_service import ExtensionDataExchangeService


class ExtensionDataExchangeServiceTests(unittest.TestCase):
    def test_csv_and_json_parse_to_object_rows(self):
        csv_rows = ExtensionDataExchangeService.parse_bytes("sample.csv", "代码,分数\n600519,1.2\n".encode())
        json_rows = ExtensionDataExchangeService.parse_bytes("sample.json", json.dumps([{"代码": "600519", "分数": 1.2}], ensure_ascii=False).encode())
        self.assertEqual(csv_rows["rows"], json_rows["rows"])
        self.assertEqual(["代码", "分数"], csv_rows["columns"])

    def test_xlsx_formulas_are_rejected_before_staging(self):
        book = Workbook()
        sheet = book.active
        sheet.append(["代码", "分数"])
        sheet.append(["600519", "=1+1"])
        output = io.BytesIO()
        book.save(output)
        with self.assertRaisesRegex(ValueError, "公式"):
            ExtensionDataExchangeService.parse_bytes("sample.xlsx", output.getvalue())

    def test_xlsx_export_contains_static_values_and_arial_font(self):
        payload = ExtensionDataExchangeService.export_rows("xlsx", ["代码", "分数"], [{"代码": "600519", "分数": 1.2}])
        book = load_workbook(io.BytesIO(payload), data_only=False)
        sheet = book.active
        self.assertEqual("代码", sheet["A1"].value)
        self.assertEqual("600519", sheet["A2"].value)
        self.assertEqual("Arial", sheet["A1"].font.name)
        self.assertFalse(any(str(cell.value).startswith("=") for row in sheet.iter_rows() for cell in row if cell.value is not None))

    def test_unsupported_extension_is_rejected(self):
        with self.assertRaisesRegex(ValueError, "仅支持"):
            ExtensionDataExchangeService.parse_bytes("sample.txt", b"hello")

    def test_http_url_requires_https_exact_allowlist_and_public_dns(self):
        with patch("app.services.extension_data_exchange_service.socket.getaddrinfo", return_value=[(None, None, None, None, ("8.8.8.8", 443))]):
            self.assertEqual("https://data.example.com/feed.csv", ExtensionDataExchangeService.validate_http_url("https://data.example.com/feed.csv", ["data.example.com"]))
        with self.assertRaisesRegex(ValueError, "HTTPS"):
            ExtensionDataExchangeService.validate_http_url("http://data.example.com/feed.csv", ["data.example.com"])

    def test_http_url_rejects_private_dns_even_when_host_is_allowlisted(self):
        with patch("app.services.extension_data_exchange_service.socket.getaddrinfo", return_value=[(None, None, None, None, ("127.0.0.1", 443))]):
            with self.assertRaisesRegex(ValueError, "非公网"):
                ExtensionDataExchangeService.validate_http_url("https://data.example.com/feed.csv", ["data.example.com"])


if __name__ == "__main__":
    unittest.main()
